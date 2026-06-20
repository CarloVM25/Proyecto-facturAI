from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def export_to_excel(invoice_data: dict, output_path: str) -> str:
    """Export invoice data to a two-sheet Excel file and return the output path."""
    wb = openpyxl.Workbook()

    _build_resumen_sheet(wb.active, invoice_data)
    _build_items_sheet(wb.create_sheet("Items"), invoice_data.get("descripcion_items") or [])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _build_resumen_sheet(ws, data: dict) -> None:
    ws.title = "Resumen"

    rows = [
        ("Tipo de Documento",     data.get("tipo_documento")),
        ("Serie y Número",        data.get("serie_numero")),
        ("Fecha de Emisión",      data.get("fecha_emision")),
        ("RUC Emisor",            data.get("ruc_emisor")),
        ("Razón Social Emisor",   data.get("razon_social_emisor")),
        ("RUC Receptor",          data.get("ruc_receptor")),
        ("Razón Social Receptor", data.get("razon_social_receptor")),
        ("Moneda",                data.get("moneda")),
        ("Subtotal",              data.get("subtotal")),
        ("IGV",                   data.get("igv")),
        ("Total",                 data.get("total")),
        ("Archivo",               data.get("filename")),
        ("Procesado",             data.get("processing_timestamp")),
    ]

    amount_row_indices = {9, 10, 11}  # Subtotal, IGV, Total (1-based)

    for row_idx, (label, value) in enumerate(rows, start=1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True)

        value_cell = ws.cell(row=row_idx, column=2, value=value)
        if row_idx in amount_row_indices and value is not None:
            value_cell.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42


def _build_items_sheet(ws, items: list) -> None:
    ws.title = "Items"

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    headers = ["Cantidad", "Descripción", "Precio Unitario", "Subtotal"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("cantidad"))
        ws.cell(row=row_idx, column=2, value=item.get("descripcion"))

        precio_cell = ws.cell(row=row_idx, column=3, value=item.get("precio_unitario"))
        precio_cell.number_format = '#,##0.00'

        subtotal_cell = ws.cell(row=row_idx, column=4, value=item.get("subtotal"))
        subtotal_cell.number_format = '#,##0.00'

    col_widths = [12, 50, 18, 15]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
